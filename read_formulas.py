import openpyxl

def analyze_formulas(file_path):
    """
    Analyzes an Excel file to extract unique formulas and their locations.

    Args:
        file_path (str): The path to the Excel file.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        all_formulas = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_formulas = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        formula = cell.value
                        if formula not in sheet_formulas:
                            sheet_formulas[formula] = []
                        sheet_formulas[formula].append(cell.coordinate)
            
            if sheet_formulas:
                all_formulas[sheet_name] = sheet_formulas

        if all_formulas:
            print(f"Found formulas in: {file_path}")
            for sheet_name, formulas in all_formulas.items():
                print(f"\n--- Sheet: {sheet_name} ---")
                for formula, cells in formulas.items():
                    print(f"  Formula: {formula}")
                    # Print only first few locations to avoid too much output
                    print(f"  Found in cells: {cells[:5]}{'...' if len(cells) > 5 else ''}")
        else:
            print(f"No formulas found in {file_path}")

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    analyze_formulas("SwimDataApp/Sample/IMX (NKB Points) 25-26-2.xlsx")